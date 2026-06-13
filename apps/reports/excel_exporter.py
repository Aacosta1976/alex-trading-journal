"""
Exportador Excel con tema dark profesional.
Replica el diseño del diario original: fondos oscuros, acentos de color,
super-headers combinados y coding por resultado W/L/Flat.
"""

import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Paleta de colores ─────────────────────────────────────────────────────────
C = {
    "dark":    "0D1117", "card":   "161B22", "header":  "1C2333",
    "gold":    "F0B429", "cyan":   "58A6FF", "green":   "3FB950",
    "red":     "F85149", "orange": "DB6D28", "purple":  "BC8CFF",
    "text":    "E6EDF3", "subtle": "8B949E", "border":  "30363D",
    "win_bg":  "1a3a1a", "loss_bg":"3a1a1a", "flat_bg": "2a2a10",
}

def _f(h):  return PatternFill("solid", fgColor=h)
def _fn(color=None, bold=False, size=9):
    return Font(color=color or C["text"], bold=bold, size=size, name="Calibri")
def _al(h="center", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def _bd():
    s = Side(border_style="thin", color=C["border"])
    return Border(left=s, right=s, top=s, bottom=s)

def _cell(ws, row, col, val, bg=None, fg=None, bold=False,
          size=9, h="center", wrap=False, nf=None):
    c = ws.cell(row=row, column=col, value=val)
    c.fill      = _f(bg or C["dark"])
    c.font      = _fn(fg or C["text"], bold, size)
    c.alignment = _al(h, "center", wrap)
    c.border    = _bd()
    if nf: c.number_format = nf
    return c


def export_excel(trades: list[dict], metrics: dict, username: str) -> bytes:
    """
    Genera el archivo Excel completo con 3 hojas:
    - Resumen KPIs
    - Diario de Operaciones
    - Estadísticas por Setup/Símbolo
    """
    wb = Workbook()
    wb.remove(wb.active)
    _sheet_summary(wb, metrics, username)
    _sheet_journal(wb, trades)
    _sheet_stats(wb, metrics)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ════════════════════════════════════════════════════════════════════════════
# HOJA 1 — RESUMEN
# ════════════════════════════════════════════════════════════════════════════
def _sheet_summary(wb, metrics, username):
    ws = wb.create_sheet("📊 Resumen")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = C["gold"]

    for r in range(1, 35):
        for c in range(1, 12):
            ws.cell(r, c).fill = _f(C["dark"])

    for c, w in {1:2, 2:28, 3:20, 4:2, 5:28, 6:20, 7:2}.items():
        ws.column_dimensions[get_column_letter(c)].width = w
    for r in range(1, 35):
        ws.row_dimensions[r].height = 18

    # Título
    ws.merge_cells("B2:F2")
    c = ws.cell(2, 2, f"⚡  TRADING JOURNAL — {username.upper()} — {datetime.now().strftime('%B %Y').upper()}")
    c.fill = _f(C["card"]); c.font = _fn(C["gold"], True, 14); c.alignment = _al("center")
    ws.row_dimensions[2].height = 30
    for col in range(2, 7): ws.cell(3, col).fill = _f(C["gold"])
    ws.row_dimensions[3].height = 3

    gross = metrics.get("gross_pnl", 0)

    # KPIs izquierda
    ws.merge_cells("B5:C5")
    _cell(ws,5,2,"RENDIMIENTO",bg=C["header"],fg=C["cyan"],bold=True,size=10,h="left")
    ws.row_dimensions[5].height = 22

    left_kpis = [
        ("Total Operaciones",    str(metrics.get("total_trades", 0)),          C["cyan"]),
        ("Win / Loss / Flat",    f"{metrics.get('total_wins',0)} / {metrics.get('total_losses',0)} / {metrics.get('total_flats',0)}", C["text"]),
        ("Win Rate",             f"{metrics.get('win_rate',0):.1f}%",          C["cyan"]),
        ("Profit Factor",        f"{metrics.get('profit_factor',0):.2f}×",     C["purple"]),
        ("Avg Risk:Reward",      f"{metrics.get('avg_rr',0):.2f}R",            C["cyan"]),
        ("Sharpe Ratio",         f"{metrics.get('sharpe_ratio',0):.2f}",       C["gold"]),
        ("Max Racha Wins",       str(metrics.get("max_win_streak", 0)),        C["green"]),
        ("Max Racha Losses",     str(metrics.get("max_loss_streak", 0)),       C["red"]),
        ("Expectancy",           f"${metrics.get('expectancy',0):.4f}",        C["cyan"]),
    ]
    for i, (lbl, val, fg) in enumerate(left_kpis):
        bg = C["card"] if i % 2 == 0 else C["dark"]
        _cell(ws, 6+i, 2, lbl, bg=bg, fg=C["subtle"], h="left")
        _cell(ws, 6+i, 3, val, bg=bg, fg=fg, bold=True)

    # KPIs derecha
    ws.merge_cells("E5:F5")
    _cell(ws,5,5,"BALANCE & P&L",bg=C["header"],fg=C["gold"],bold=True,size=10,h="left")

    right_kpis = [
        ("Balance Inicial",  f"${metrics.get('initial_balance',0):,.2f}",                    C["text"]),
        ("Balance Final",    f"${metrics.get('final_balance',0):,.2f}",                      C["green"] if gross >= 0 else C["red"]),
        ("P&L Neto",         f"${gross:+,.2f}",                                              C["green"] if gross >= 0 else C["red"]),
        ("ROE",              f"{metrics.get('roe',0):+.2f}%",                                C["green"] if metrics.get("roe",0) >= 0 else C["red"]),
        ("Avg Win ($)",      f"${metrics.get('avg_win',0):.2f}",                             C["green"]),
        ("Avg Loss ($)",     f"${metrics.get('avg_loss',0):.2f}",                            C["red"]),
        ("Mejor Trade",      f"${metrics.get('best_trade',0):.2f}",                          C["green"]),
        ("Peor Trade",       f"${metrics.get('worst_trade',0):.2f}",                         C["red"]),
        ("Max Drawdown",     f"${metrics.get('max_drawdown',0):.2f} ({metrics.get('max_drawdown_pct',0):.1f}%)", C["red"]),
    ]
    for i, (lbl, val, fg) in enumerate(right_kpis):
        bg = C["card"] if i % 2 == 0 else C["dark"]
        _cell(ws, 6+i, 5, lbl, bg=bg, fg=C["subtle"], h="left")
        _cell(ws, 6+i, 6, val, bg=bg, fg=fg, bold=True)


# ════════════════════════════════════════════════════════════════════════════
# HOJA 2 — DIARIO DE OPERACIONES
# ════════════════════════════════════════════════════════════════════════════
def _sheet_journal(wb, trades):
    ws = wb.create_sheet("📋 Diario")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = C["cyan"]

    col_defs = [
        ("#",         5,  C["subtle"]),  ("Fecha",    12, C["subtle"]),
        ("Símbolo",  11,  C["gold"]),    ("Pos.",      8, C["text"]),
        ("Setup",     7,  C["purple"]),  ("TF",        7, C["subtle"]),
        ("Risk%",     8,  C["cyan"]),    ("Pips SL",   9, C["text"]),
        ("Size K",    9,  C["text"]),    ("R:R Init",  9, C["cyan"]),
        ("R:R Real",  9,  C["cyan"]),    ("P&L Real", 12, C["green"]),
        ("W/L",       7,  C["green"]),   ("Acum $",   13, C["gold"]),
        ("Error",    24,  C["orange"]),  ("Gestión",  20, C["subtle"]),
        ("Notas",    30,  C["subtle"]),
    ]

    for r in range(1, len(trades) + 8):
        for c in range(1, len(col_defs) + 1):
            ws.cell(r, c).fill = _f(C["dark"])
    for i, (_, w, _) in enumerate(col_defs, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Título
    ws.merge_cells(f"A1:{get_column_letter(len(col_defs))}1")
    t = ws.cell(1, 1, "📋  DIARIO DE OPERACIONES")
    t.fill = _f(C["card"]); t.font = _fn(C["gold"], True, 13); t.alignment = _al("center")
    ws.row_dimensions[1].height = 26

    # Cabecera
    for i, (hdr, _, fg) in enumerate(col_defs, 1):
        _cell(ws, 2, i, hdr, bg=C["header"], fg=fg, bold=True, size=8)
    ws.row_dimensions[2].height = 20
    ws.freeze_panes = "A3"

    # Datos
    for idx, t in enumerate(trades):
        row = 3 + idx
        bg  = C["card"] if idx % 2 == 0 else C["dark"]
        wl  = t.get("result", "")
        pnl = float(t.get("pnl_real") or 0)
        acum= float(t.get("profit_cumul") or 0)
        risk= t.get("risk_pct")
        ri  = t.get("ratio_initial")
        rr  = t.get("ratio_real")
        sz  = t.get("size_k")
        pos = t.get("position", "")

        _cell(ws, row,  1, idx+1,                                         bg=bg, fg=C["subtle"])
        _cell(ws, row,  2, str(t.get("trade_date","")),                   bg=bg, fg=C["subtle"])
        _cell(ws, row,  3, t.get("symbol",""),                            bg=bg, fg=C["gold"],   bold=True)
        _cell(ws, row,  4, pos,                                           bg=bg, fg=C["green"] if pos=="Long" else C["red"], bold=True)
        _cell(ws, row,  5, t.get("setup",""),                             bg=bg, fg=C["purple"], bold=True)
        _cell(ws, row,  6, t.get("timeframe",""),                         bg=bg, fg=C["subtle"])
        _cell(ws, row,  7, f"{float(risk)*100:.1f}%" if risk else "",     bg=bg, fg=C["cyan"])
        _cell(ws, row,  8, t.get("pips_sl",""),                           bg=bg, fg=C["text"])
        _cell(ws, row,  9, f"{float(sz):.3f}" if sz else "",              bg=bg, fg=C["text"])
        _cell(ws, row, 10, f"{float(ri):.2f}R" if ri else "",             bg=bg, fg=C["cyan"])
        _cell(ws, row, 11, f"{float(rr):.4f}R" if rr else "",             bg=bg, fg=C["cyan"])
        _cell(ws, row, 12, f"${pnl:+.2f}",                               bg=bg, fg=C["green"] if pnl>0 else C["red"] if pnl<0 else C["orange"], bold=True)
        # W/L con fondo de color
        wl_bg = C["win_bg"] if wl=="W" else (C["loss_bg"] if wl=="L" else C["flat_bg"])
        wl_fg = C["green"]  if wl=="W" else (C["red"]     if wl=="L" else C["orange"])
        _cell(ws, row, 13, wl,                                            bg=wl_bg, fg=wl_fg, bold=True)
        _cell(ws, row, 14, f"${acum:+.2f}",                              bg=bg, fg=C["green"] if acum>=0 else C["red"], bold=True)
        _cell(ws, row, 15, t.get("error_type",""),                        bg=bg, fg=C["orange"], h="left")
        _cell(ws, row, 16, t.get("trade_mgmt",""),                        bg=bg, fg=C["subtle"], h="left")
        _cell(ws, row, 17, t.get("notes",""),                             bg=bg, fg=C["subtle"], h="left", wrap=True)
        ws.row_dimensions[row].height = 16


# ════════════════════════════════════════════════════════════════════════════
# HOJA 3 — ESTADÍSTICAS
# ════════════════════════════════════════════════════════════════════════════
def _sheet_stats(wb, metrics):
    ws = wb.create_sheet("📈 Estadísticas")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = C["green"]

    for r in range(1, 40):
        for c in range(1, 12):
            ws.cell(r, c).fill = _f(C["dark"])

    for c, w in {1:2, 2:26, 3:10, 4:10, 5:10, 6:12, 7:2, 8:26, 9:10}.items():
        ws.column_dimensions[get_column_letter(c)].width = w

    ws.merge_cells("B2:F2")
    h = ws.cell(2, 2, "📈  ESTADÍSTICAS POR SETUP Y SÍMBOLO")
    h.fill = _f(C["card"]); h.font = _fn(C["gold"], True, 13); h.alignment = _al("center")
    ws.row_dimensions[2].height = 26
    for col in range(2, 7): ws.cell(3, col).fill = _f(C["gold"])
    ws.row_dimensions[3].height = 3

    # Setup stats
    ws.merge_cells("B5:F5")
    _cell(ws,5,2,"PERFORMANCE POR SETUP",bg=C["header"],fg=C["cyan"],bold=True,size=10,h="left")
    for i, h in enumerate(["Setup","Trades","Wins","Losses","Win Rate","P&L"]):
        _cell(ws, 6, 2+i, h, bg=C["header"], fg=C["cyan"], bold=True, size=8)

    for i, (k, v) in enumerate(metrics.get("setup_stats",{}).items()):
        bg = C["card"] if i%2==0 else C["dark"]
        _cell(ws,7+i,2,k,         bg=bg,fg=C["purple"],bold=True)
        _cell(ws,7+i,3,v["total"],bg=bg,fg=C["text"])
        _cell(ws,7+i,4,v["wins"],  bg=bg,fg=C["green"],bold=True)
        _cell(ws,7+i,5,v["losses"],bg=bg,fg=C["red"],  bold=True)
        _cell(ws,7+i,6,f"{v['win_rate']:.1f}%",
              bg=bg, fg=C["green"] if v["win_rate"]>=60 else (C["red"] if v["win_rate"]<40 else C["orange"]),bold=True)
        pnl_v = v["pnl"]
        _cell(ws,7+i,7,f"${pnl_v:+.2f}",bg=bg,fg=C["green"] if pnl_v>=0 else C["red"],bold=True)

    # Symbol stats
    row_s = 7 + len(metrics.get("setup_stats",{})) + 2
    ws.merge_cells(f"B{row_s}:G{row_s}")
    _cell(ws,row_s,2,"PERFORMANCE POR SÍMBOLO",bg=C["header"],fg=C["gold"],bold=True,size=10,h="left")
    for i, h in enumerate(["Símbolo","Trades","Wins","Losses","Win Rate","P&L"]):
        _cell(ws,row_s+1,2+i,h,bg=C["header"],fg=C["gold"],bold=True,size=8)

    for i, (k, v) in enumerate(metrics.get("symbol_stats",{}).items()):
        bg = C["card"] if i%2==0 else C["dark"]
        _cell(ws,row_s+2+i,2,k,          bg=bg,fg=C["gold"],bold=True)
        _cell(ws,row_s+2+i,3,v["total"], bg=bg,fg=C["text"])
        _cell(ws,row_s+2+i,4,v["wins"],  bg=bg,fg=C["green"],bold=True)
        _cell(ws,row_s+2+i,5,v["losses"],bg=bg,fg=C["red"],  bold=True)
        _cell(ws,row_s+2+i,6,f"{v['win_rate']:.1f}%",
              bg=bg,fg=C["green"] if v["win_rate"]>=60 else C["red"],bold=True)
        pnl_v = v["pnl"]
        _cell(ws,row_s+2+i,7,f"${pnl_v:+.2f}",bg=bg,fg=C["green"] if pnl_v>=0 else C["red"],bold=True)
